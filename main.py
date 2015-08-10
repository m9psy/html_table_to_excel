# Shebang
# coding: utf-8
"""
Transforms raw html to xlsx file, considering colspans, rowspans and styles.
"""

__author__ = 'Dmitriy Emelianov'

import openpyxl
from bs4 import BeautifulSoup
from docx import Document


def html_to_workbook(html_text):
    """
    Transforms raw html table to Excel workbook
    :param html_text: Raw html table.
    :return: openpyxl.Workbook object
    """
    tables = BeautifulSoup(html_text).findAll('table')
    wb = openpyxl.Workbook()
    for table_index in range(len(tables)):
        if table_index == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = "Table_" + str(table_index)
        rows = tables[table_index].findAll('tr')
        max_columns = max([len(rows[row_index].findAll(['th', 'td'])) for row_index in range(len(rows))])
        column_shifter = [[None for elem_index in range(max_columns)] for row_index in range(len(rows))]
        column_seeker = [0 for row_index in range(len(rows))]
        for row_index in range(len(rows)):
            cols = rows[row_index].findAll(['th', 'td'])
            for col_index in range(len(cols)):
                rowspan = int(cols[col_index].get('rowspan', 1))
                colspan = int(cols[col_index].get('colspan', 1))
                if rowspan == 0: rowspan = 1
                if colspan == 0: colspan = 1
                for rspan in range(rowspan):
                    if rspan == 0:
                        column_shifter[row_index + rspan][column_seeker[row_index + rspan]] = col_index
                    # Shift columns only on right side, do not touch columns on left side, they are not affected by colspan
                    if column_seeker[row_index] - colspan <= column_seeker[row_index + rspan]:
                        column_seeker[row_index + rspan] += colspan
                if rowspan != 1 or colspan != 1:
                    ws.merge_cells(start_row=row_index + 1, start_column=column_seeker[row_index] - colspan + 1,
                                   end_row=row_index + rowspan, end_column=column_seeker[row_index])
                try:
                    result = float(cols[col_index].getText())
                    if result.is_integer():
                        result = int(cols[col_index].getText())
                except ValueError:
                    result = cols[col_index].getText()
                finally:
                    excel_index = cni(column_shifter[row_index].index(col_index), row_index)
                    ws[excel_index] = result
                    ws[excel_index].alignment = Alignment(horizontal="center", vertical="center")
    return wb


def html_to_docx(html_text):
    tables = BeautifulSoup(html_text).findAll('table')
    document = Document()
    for table_index in range(len(tables)):
        rows = tables[table_index].findAll('tr')
        max_columns = max([len(rows[row_index].findAll(['th', 'td'])) for row_index in range(len(rows))])
        column_shifter = [[None for elem_index in range(max_columns)] for row_index in range(len(rows))]
        column_seeker = [0 for row_index in range(len(rows))]
        table = document.add_table(rows=len(rows), cols=max_columns)
        table.style = 'TableGrid'
        for row_index in range(len(rows)):
            cols = rows[row_index].findAll(['th', 'td'])
            for col_index in range(len(cols)):
                rowspan = int(cols[col_index].get('rowspan', 1))
                colspan = int(cols[col_index].get('colspan', 1))
                if rowspan == 0: rowspan = 1
                if colspan == 0: colspan = 1
                for rspan in range(rowspan):
                    if rspan == 0:
                        column_shifter[row_index + rspan][column_seeker[row_index + rspan]] = col_index
                    # Shift columns only on right side, do not touch columns on left side, they are not affected by colspan
                    if column_seeker[row_index] - colspan <= column_seeker[row_index + rspan]:
                        column_seeker[row_index + rspan] += colspan
                if rowspan != 1 or colspan != 1:
                    top_left = table.cell(row_index, column_seeker[row_index] - colspan)
                    bottom_right = table.cell(row_index + rowspan - 1, column_seeker[row_index] - 1)
                    current_cell = top_left.merge(bottom_right)
                else:
                    current_cell = table.cell(row_index, column_shifter[row_index].index(col_index))
                try:
                    result = float(cols[col_index].getText())
                    if result.is_integer():
                        result = int(cols[col_index].getText())
                except ValueError:
                    result = cols[col_index].getText()
                finally:
                    current_cell.text = str(result)
    return document